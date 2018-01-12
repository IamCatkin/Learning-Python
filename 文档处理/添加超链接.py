#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/1/12 10:09
# @Author  : Catkin
# @File    : 添加超链接.py
from openpyxl import Workbook
wb = Workbook()
sheet1 = wb.active
sheet1.cell(row=2,column=1).value='P31749'
sheet1.cell(row=2,column=2).value='1H10'
sheet1.cell(row=2,column=3).value="http://www.rcsb.org/pdb/explore/explore.do?structureId=" + sheet1.cell(row=2,column=2).value
links = "=HYPERLINK(\""+sheet1.cell(row=2,column=3).value+'\",\"'+sheet1.cell(row=2,column=2).value+'\")'
sheet1.cell(row=2,column=4).value=links
wb.save('test.xlsx')
