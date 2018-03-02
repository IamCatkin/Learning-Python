#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/2/28 19:19
# @Author  : Catkin
# @Website : blog.catkin.moe
# @File    : Food Additives.py
import requests
import re
import pandas
from time import sleep
from openpyxl import load_workbook

wb = load_workbook('Food Additives.xlsx')
sheet = wb['Sheet1']
line = 2

page = requests.get(url="http://apps.who.int/food-additives-contaminants-jecfa-database/search.aspx?fcc=1",timeout=50.0).text
patterns = re.compile('<div id="SearchResultItem"><a href=chemical.aspx\?chemID=(.*?)>')
lists =  re.findall(patterns,page)

for item in lists:
    try:
        sheet.cell(row=line,column=1).value=item
        chem = requests.get(url="http://apps.who.int/food-additives-contaminants-jecfa-database/chemical.aspx?chemID="+item,timeout=50.0).text
        data = pandas.read_html(chem)
        if item in ["5865","5866","5868"]:
            pass
        else:
            if len(data)<2:
                table_1 = data[0].set_index(0)
                if 'Chemical Names:' in table_1.index:
                    sheet.cell(row=line,column=2).value=table_1.loc['Chemical Names:',1]
                else:
                    sheet.cell(row=line,column=2).value=""
                if 'CAS number:' in table_1.index:
                    sheet.cell(row=line,column=3).value=table_1.loc['CAS number:',1]
                else:
                    sheet.cell(row=line,column=3).value=""        
                if 'Functional Class:' in table_1.index:
                    sheet.cell(row=line,column=4).value=table_1.loc['Functional Class:',1]
                else:
                    sheet.cell(row=line,column=4).value=""   
            elif len(data)==2:
                table_1 = data[0].set_index(0)
                table_2 = data[1].set_index(0)
                if 'Chemical Names:' in table_1.index:
                    sheet.cell(row=line,column=2).value=table_1.loc['Chemical Names:',1]
                else:
                    sheet.cell(row=line,column=2).value=""
                if 'CAS number:' in table_1.index:
                    sheet.cell(row=line,column=3).value=table_1.loc['CAS number:',1]
                else:
                    sheet.cell(row=line,column=3).value=""        
                if 'Functional Class:' in table_1.index:
                    sheet.cell(row=line,column=4).value=table_1.loc['Functional Class:',1]
                else:
                    sheet.cell(row=line,column=4).value=""   
                if 'Evaluation year:' in table_2.index:
                    sheet.cell(row=line,column=5).value=table_2.loc['Evaluation year:',1]
                else:
                    sheet.cell(row=line,column=5).value=""        
                if 'ADI:' in table_2.index:
                    sheet.cell(row=line,column=6).value=table_2.loc['ADI:',1]
                elif 'Tolerable Intake:' in table_2.index:
                    sheet.cell(row=line,column=12).value=table_2.loc['Tolerable Intake:',1]
                else:
                    sheet.cell(row=line,column=6).value=""        
                if 'Specs Code:' in table_2.index:
                    sheet.cell(row=line,column=7).value=table_2.loc['Specs Code:',1]
                else:
                    sheet.cell(row=line,column=7).value=""     
                if 'Report:' in table_2.index:
                    sheet.cell(row=line,column=8).value=table_2.loc['Report:',1]
                else:
                    sheet.cell(row=line,column=8).value=""
                if 'Specification:' in table_2.index:
                    sheet.cell(row=line,column=9).value=table_2.loc['Specification:',1]
                else:
                    sheet.cell(row=line,column=9).value=""     
                if 'Previous Years:' in table_2.index:
                    sheet.cell(row=line,column=10).value=table_2.loc['Previous Years:',1]
                else:
                    sheet.cell(row=line,column=10).value=""
                if 'Tox Monograph:' in table_2.index:
                    sheet.cell(row=line,column=11).value=table_2.loc['Tox Monograph:',1]
                else:
                    sheet.cell(row=line,column=11).value=""    
            elif len(data)>2:
                table_1 = data[0].set_index(0)
                table_2 = data[1].set_index(0)
                table_3 = data[2].set_index(0)
                if 'Chemical Names:' in table_1.index:
                    sheet.cell(row=line,column=2).value=table_1.loc['Chemical Names:',1]
                else:
                    sheet.cell(row=line,column=2).value=""
                if 'CAS number:' in table_1.index:
                    sheet.cell(row=line,column=3).value=table_1.loc['CAS number:',1]
                else:
                    sheet.cell(row=line,column=3).value=""        
                if 'Functional Class:' in table_1.index:
                    sheet.cell(row=line,column=4).value=table_1.loc['Functional Class:',1]
                else:
                    sheet.cell(row=line,column=4).value=""
                if 'Evaluation year:' in table_2.index:
                    sheet.cell(row=line,column=5).value=table_2.loc['Evaluation year:',1]
                else:
                    sheet.cell(row=line,column=5).value=""        
                if 'ADI:' in table_2.index:
                    sheet.cell(row=line,column=6).value=table_2.loc['ADI:',1]
                elif 'Tolerable Intake:' in table_2.index:
                    sheet.cell(row=line,column=12).value=table_2.loc['Tolerable Intake:',1]
                else:
                    sheet.cell(row=line,column=6).value=""    
                if 'Specs Code:' in table_2.index:
                    sheet.cell(row=line,column=7).value=table_2.loc['Specs Code:',1]
                else:
                    sheet.cell(row=line,column=7).value=""     
                if 'Report:' in table_2.index:
                    sheet.cell(row=line,column=8).value=table_2.loc['Report:',1]
                else:
                    sheet.cell(row=line,column=8).value=""
                if 'Specification:' in table_2.index:
                    sheet.cell(row=line,column=9).value=table_2.loc['Specification:',1]
                else:
                    sheet.cell(row=line,column=9).value=""    
                if 'Previous Years:' in table_2.index:
                    sheet.cell(row=line,column=10).value=table_2.loc['Previous Years:',1]
                elif 'Previous Years:' in table_3.index:
                    sheet.cell(row=line,column=10).value=table_3.loc['Previous Years:',1]
                else:
                    sheet.cell(row=line,column=10).value=""
                if 'Tox Monograph:' in table_2.index:
                    sheet.cell(row=line,column=11).value=table_2.loc['Tox Monograph:',1]
                else:
                    sheet.cell(row=line,column=11).value=""    
            sleep(0.5)
        line += 1
    finally:
        wb.save('results.xlsx')

