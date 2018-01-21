#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/1/12 15:47
# @Author  : Catkin
# @File    : 查找配体.py
import requests
import re
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.poolmanager import PoolManager
import ssl
from openpyxl import load_workbook
from time import sleep

class MyAdapter(HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(num_pools=connections,maxsize=maxsize,block=block,ssl_version=ssl.PROTOCOL_TLSv1)

s = requests.Session()
s.mount('https://', MyAdapter())
headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 6.0; WOW64; rv:24.0) Gecko/20100101 Firefox/24.0' }

wb = load_workbook('Uniprot_PDB.xlsx')
sheet = wb['Sheet1']
try:
    for i in range(2,18113):
        uniprot = sheet.cell(row=i,column=1).value
        pdb = sheet.cell(row=i,column=2).value
        urls = "https://www.rcsb.org/structure/"+str(pdb)
        page = s.get(url=urls,headers=headers,timeout=50.0).text
        if "LigandsTable" in page:
            patterns = re.compile('https://files.rcsb.org/ligands/download/(.*?).cif')
            liagands = re.findall(patterns,page)
            if len(liagands) != 0:
                for j in range(len(liagands)):
                    print(liagands[j])
                    filename = ".\\ligand\\" + uniprot + '_' + str(pdb) + '_' + liagands[j] + '.sdf'
                    downloadlinks = "https://www.rcsb.org/pdb/download/downloadLigandFiles.do?ligandIdList=" + liagands[j] + "&structIdList=" + str(pdb) + "&instanceType=all&excludeUnobserved=false&includeHydrogens=false"
                    files = s.get(url=downloadlinks,headers=headers,timeout=50.0)
                    if files.status_code == 200:
                        with open(filename, 'wb') as d:
                            d.write(files.content)
                    else:
                        print('Downloaderror')
                        sheet.cell(row=i, column=4).value = 'Error'
                        with open('error.txt', 'a+') as f:
                            f.write(str(i) + "下载错误" + "\n")
                            f.close()
                    sheet.cell(row=i,column=4+j).value=liagands[j]
            else:
                sheet.cell(row=i, column=4).value = 'Fault'
                print('Fault')
        elif "Page not found" in page:
            print('Error')
            sheet.cell(row=i, column=4).value = 'Error'
            with open('error.txt','a+') as f:
                f.write(str(i)+"\n")
                f.close()
        elif "LigandsTable" not in page:
            sheet.cell(row=i,column=4).value='None'
            print('None')
        sleep(0.6)
finally:
    wb.save('results.xlsx')

