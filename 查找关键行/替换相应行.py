#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/12/27 11:09
# @Author  : Catkin
# @File    : 替换相应行.py
from openpyxl import Workbook
from openpyxl import load_workbook

data1 = load_workbook(r'11.xlsx')
data2 = load_workbook(r'22.xlsx')
sheet1 = data1['11']
sheet2 = data2['22']
# print(sheet1.cell(row=1,column=1).value)

count =0
for i in range(1,7974):
    for j in range(1,642):
        if sheet1.cell(row=i,column=3).value == sheet2.cell(row=j,column=2).value:
            sheet2.cell(row=j, column=3).value = sheet1.cell(row=i,column=1).value
            count += 1
print(count)
data2.save('55.xlsx')
