#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/1/7 16:31
# @Author  : Catkin
# @File    : predict.py
import os
import pandas as pd
from openpyxl import Workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split 

#########   遍历目录下所有文件 ########
series = ["series1","series2","all data","literature data"]
path = ".\\txt\\"

#########   读取文件系列1   ############
for s in series:
    if s == "series1":
        files_1 = os.listdir(path+s)  
        files = []
        for f in files_1:
            if f != 'hepdata0_1038_2d.txt':
                files.append(f)

        for file in files:
            cycle = 0
            filename = ".\\results\\"+s+"\\"+file[:-4]
            wb1 = Workbook()
            sheet1 = wb1.active
            while cycle < 20:
                r1 = 2 * cycle + 2
                r2 = r1 + 1
                data = pd.read_table(path+s+"\\"+file,delimiter='\t')
                positive = data.iloc[0:len(data.index),3:len(data.columns)]
                data_0 = pd.read_table(path+s+"\\"+"hepdata0_1038_2d.txt",delimiter='\t')
                negative = data_0.iloc[0:len(data_0.index),3:len(data.columns)]
                random_negative = negative.sample(n=len(data.index))
                complain = pd.concat([positive,random_negative])
                X = complain.iloc[:,1:len(complain.columns)]
                y = complain.iloc[:,0]
                positive_label = int(complain.iloc[0,0])
                pre_data_0 = pd.read_table(path+'database_test.txt',delimiter='\t',na_values='NaN')
                pre_data_0 = pre_data_0.loc[pre_data_0["lable"]==positive_label]
                pre_data = pre_data_0.iloc[:,4:len(pre_data_0.columns)]
        
#########  划分训练集与测试集 #########
                X_train,X_,y_train,y_ = train_test_split(X,y,test_size=0.0)

#########   使用随机森林算法进行预测 ########
                rf = RandomForestClassifier(n_estimators=500).fit(X_train,y_train)
                answer_rf = rf.predict(pre_data)

#########   评价结果　　　#################
                chance = rf.predict_proba(pre_data)[:,1]

########## 存数据   ##############
                j = 2
                sheet1.cell(row=1,column=1).value = "SMILES"
                sheet1.cell(row=1,column=r1).value = "Pre"
                sheet1.cell(row=1,column=r2).value = "pos_chance"
                for r,c in zip(answer_rf,chance):
                    sheet1.cell(row=j,column=1).value = pre_data_0.iloc[j-2,1]
                    sheet1.cell(row=j,column=r1).value = r
                    sheet1.cell(row=j,column=r2).value = c
                    j += 1
                cycle += 1
            wb1.save(filename+".xlsx")
                
#########   读取文件系列2   ############
    elif s == "series2":
        files_2 = os.listdir(path+s)  
        files = []
        for f in files_2:
            if f != 'hepdata0_1038_2d.txt':
                files.append(f)

        for file in files:
            cycle = 0
            filename = ".\\results\\"+s+"\\"+file[:-4]
            wb2 = Workbook()
            sheet2 = wb2.active
            while cycle < 20:
                r1 = 2 * cycle + 2
                r2 = r1 + 1
                data = pd.read_table(path+s+"\\"+file,delimiter='\t')
                positive = data.iloc[0:len(data.index),3:len(data.columns)]
                data_0 = pd.read_table(path+s+"\\"+"hepdata0_1038_2d.txt",delimiter='\t')
                negative = data_0.iloc[0:len(data_0.index),3:len(data.columns)]
                random_negative = negative.sample(n=len(data.index))
                complain = pd.concat([positive,random_negative])
                X = complain.iloc[:,1:len(complain.columns)]
                y = complain.iloc[:,0]
                positive_label = int(complain.iloc[0,0])
                pre_data_0 = pd.read_table(path+'database_test.txt',delimiter='\t',na_values='NaN')
                pre_data_0 = pre_data_0.loc[pre_data_0["lable"]==positive_label]
                pre_data = pre_data_0.iloc[:,4:len(pre_data_0.columns)]
        
#########  划分训练集与测试集 #########
                X_train,X_,y_train,y_ = train_test_split(X,y,test_size=0.0)

#########   使用随机森林算法进行预测 ########
                rf = RandomForestClassifier(n_estimators=500).fit(X_train,y_train)
                answer_rf = rf.predict(pre_data)

#########   评价结果　　　#################
                chance = rf.predict_proba(pre_data)[:,1]

########## 存数据   ##############
                j = 2
                sheet2.cell(row=1,column=1).value = "SMILES"
                sheet2.cell(row=1,column=r1).value = "Pre"
                sheet2.cell(row=1,column=r2).value = "pos_chance"
                for r,c in zip(answer_rf,chance):
                    sheet2.cell(row=j,column=1).value = pre_data_0.iloc[j-2,1]
                    sheet2.cell(row=j,column=r1).value = r
                    sheet2.cell(row=j,column=r2).value = c
                    j += 1
                cycle += 1
            wb2.save(filename+".xlsx")
                
#########   读取文件系列3   ############
    elif s == "all data":
        cycle = 0
        file = "all_data.txt"
        filename = filename = ".\\results\\"+s+"\\"+"all_data"
        wb3 = Workbook()
        sheet3 = wb3.active
        while cycle < 20:
            r1 = 2 * cycle + 2
            r2 = r1 + 1
            data = pd.read_table(path+s+"\\"+file,delimiter='\t')
            X = data.iloc[:,4:len(data.columns)]
            y = data.iloc[:,3]
            pre_data_0 = pd.read_table(path+'database_test.txt',delimiter='\t',na_values='NaN')
            pre_data = pre_data_0.iloc[:,4:len(pre_data_0.columns)]
            
#########  划分训练集与测试集 #########
            X_train,X_,y_train,y_ = train_test_split(X,y,test_size=0.0)

#########   使用随机森林算法进行预测 ########
            rf = RandomForestClassifier(n_estimators=500).fit(X_train,y_train)
            answer_rf = rf.predict(pre_data)

#########   评价结果　　　#################
            chance = rf.predict_proba(pre_data)[:,1]

########## 存数据   ##############
            j = 2
            sheet3.cell(row=1,column=1).value = "SMILES"
            sheet3.cell(row=1,column=r1).value = "Pre"
            sheet3.cell(row=1,column=r2).value = "pos_chance"
            for r,c in zip(answer_rf,chance):
                sheet3.cell(row=j,column=1).value = pre_data_0.iloc[j-2,1]
                sheet3.cell(row=j,column=r1).value = r
                sheet3.cell(row=j,column=r2).value = c
                j += 1
            cycle += 1
        wb3.save(filename+".xlsx")    
            
#########   读取文件系列4   ############
    else:        
        cycle = 0
        file = "literdata_2d.txt"
        filename = filename = ".\\results\\"+s+"\\"+"literdata_2d"
        wb4 = Workbook()
        sheet4 = wb4.active
        while cycle < 20:       
            r1 = 2 * cycle + 2
            r2 = r1 + 1
            data = pd.read_table(path+s+"\\"+file,delimiter='\t')
            X = data.iloc[:,4:len(data.columns)]
            y = data.iloc[:,3]
            pre_data_0 = pd.read_table(path+'database_test.txt',delimiter='\t',na_values='NaN')
            pre_data = pre_data_0.iloc[:,4:len(pre_data_0.columns)]
            
#########  划分训练集与测试集 #########
            X_train,X_,y_train,y_ = train_test_split(X,y,test_size=0.0)

#########   使用随机森林算法进行预测 ########
            rf = RandomForestClassifier(n_estimators=500).fit(X_train,y_train)
            answer_rf = rf.predict(pre_data)

#########   评价结果　　　#################
            chance = rf.predict_proba(pre_data)[:,1]

########## 存数据   ##############
            j = 2
            sheet4.cell(row=1,column=1).value = "SMILES"
            sheet4.cell(row=1,column=r1).value = "Pre"
            sheet4.cell(row=1,column=r2).value = "pos_chance"
            for r,c in zip(answer_rf,chance):
                sheet4.cell(row=j,column=1).value = pre_data_0.iloc[j-2,1]
                sheet4.cell(row=j,column=r1).value = r
                sheet4.cell(row=j,column=r2).value = c
                j += 1
            cycle += 1
        wb4.save(filename+".xlsx")              